from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .formula_utils import (
    detect_formula_risk,
    excel_formula_to_js_hint,
    extract_references,
    normalize_formula,
)


@dataclass
class CellDocument:
    id: str
    workbook: str
    sheet: str
    address: str
    row: int
    column: int
    label: str
    value: Any
    formula: str
    number_format: str
    dependencies: list[str]
    risk_level: str
    risk_reasons: list[str]
    nearby_context: dict[str, Any]
    text: str
    js_hint: str


def stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_cell_value(ws, row: int, col: int) -> Any:
    if row < 1 or col < 1:
        return None
    return ws.cell(row=row, column=col).value


def find_nearby_label(ws_formula, row: int, col: int, max_left: int = 6, max_up: int = 4) -> str:
    """Find a human label around a cell.

    Engineering spreadsheets often put labels to the left or above a formula/value.
    This heuristic prioritizes left cells, then cells above.
    """
    candidates: list[str] = []

    for offset in range(1, max_left + 1):
        v = get_cell_value(ws_formula, row, col - offset)
        if isinstance(v, str) and v.strip() and not v.strip().startswith("="):
            candidates.append(v.strip())
            break

    for offset in range(1, max_up + 1):
        v = get_cell_value(ws_formula, row - offset, col)
        if isinstance(v, str) and v.strip() and not v.strip().startswith("="):
            candidates.append(v.strip())
            break

    return " | ".join(candidates)


def get_nearby_context(ws_formula, ws_values, row: int, col: int, radius: int = 1) -> dict[str, Any]:
    context: dict[str, Any] = {}
    for r in range(max(1, row - radius), min(ws_formula.max_row, row + radius) + 1):
        for c in range(max(1, col - radius), min(ws_formula.max_column, col + radius) + 1):
            addr = f"{get_column_letter(c)}{r}"
            formula_value = ws_formula.cell(r, c).value
            computed_value = ws_values.cell(r, c).value
            if formula_value is not None or computed_value is not None:
                context[addr] = {
                    "formula_or_text": formula_value,
                    "computed_value": computed_value,
                }
    return context


def should_index_cell(value: Any, formula: str, label: str) -> bool:
    if formula:
        return True
    if label and value not in (None, ""):
        return True
    if isinstance(value, str) and len(value.strip()) >= 3:
        return True
    return False


def build_text(doc: dict[str, Any]) -> str:
    pieces = [
        f"Workbook: {doc['workbook']}",
        f"Sheet: {doc['sheet']}",
        f"Cell: {doc['address']}",
        f"Label: {doc['label']}",
        f"Value: {doc['value']}",
        f"Formula: {doc['formula']}",
        f"Dependencies: {', '.join(doc['dependencies'])}",
        f"Risk: {doc['risk_level']} {' '.join(doc['risk_reasons'])}",
    ]
    for addr, item in doc["nearby_context"].items():
        pieces.append(f"Nearby {addr}: {item}")
    return "\n".join(p for p in pieces if p and p != "None")


def extract_sheet_documents(
    workbook_path: str | Path,
    sheet_name: str = "Vibration",
    output_jsonl: str | Path | None = None,
    only_formulas: bool = False,
) -> list[CellDocument]:
    workbook_path = Path(workbook_path)
    wb_formula = load_workbook(workbook_path, data_only=False, read_only=False)
    wb_values = load_workbook(workbook_path, data_only=True, read_only=False)

    if sheet_name not in wb_formula.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb_formula.sheetnames}")

    ws_formula = wb_formula[sheet_name]
    ws_values = wb_values[sheet_name]

    docs: list[CellDocument] = []
    for row in range(1, ws_formula.max_row + 1):
        for col in range(1, ws_formula.max_column + 1):
            cell = ws_formula.cell(row, col)
            value_cell = ws_values.cell(row, col)
            raw = cell.value
            formula = raw if isinstance(raw, str) and raw.startswith("=") else ""
            value = value_cell.value
            label = find_nearby_label(ws_formula, row, col)

            if only_formulas and not formula:
                continue
            if not should_index_cell(value, formula, label):
                continue

            address = f"{get_column_letter(col)}{row}"
            dependencies = extract_references(formula)
            risk = detect_formula_risk(formula)
            nearby_context = get_nearby_context(ws_formula, ws_values, row, col, radius=1)

            temp = {
                "workbook": workbook_path.name,
                "sheet": sheet_name,
                "address": address,
                "label": label,
                "value": value,
                "formula": normalize_formula(formula),
                "dependencies": dependencies,
                "risk_level": risk.level,
                "risk_reasons": risk.reasons,
                "nearby_context": nearby_context,
            }

            doc = CellDocument(
                id=f"{workbook_path.stem}::{sheet_name}!{address}",
                workbook=workbook_path.name,
                sheet=sheet_name,
                address=address,
                row=row,
                column=col,
                label=label,
                value=value,
                formula=normalize_formula(formula),
                number_format=cell.number_format,
                dependencies=dependencies,
                risk_level=risk.level,
                risk_reasons=risk.reasons,
                nearby_context=nearby_context,
                text=build_text(temp),
                js_hint=excel_formula_to_js_hint(formula),
            )
            docs.append(doc)

    if output_jsonl:
        output_path = Path(output_jsonl)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8") as f:
            for doc in docs:
                f.write(json.dumps(asdict(doc), ensure_ascii=False, default=str) + "\n")

    return docs
