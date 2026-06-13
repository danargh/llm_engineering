from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .formula_utils import (
    detect_formula_risk,
    excel_formula_to_js_hint,
    extract_references,
    normalize_formula,
)
from .variable_naming import (
    build_variable_registry,
    formula_to_named_js,
    named_dependencies,
)


@dataclass
class CellDocument:
    id: str
    workbook: str
    sheet: str
    address: str
    row: int
    column: int
    label: str
    value: Any
    formula: str
    number_format: str
    dependencies: list[str]
    variable_name: str
    variable_role: str
    named_dependencies: dict[str, Any]
    js_formula_named: str
    risk_level: str
    risk_reasons: list[str]
    nearby_context: dict[str, Any]
    section_title: str
    section_path: list[str]
    section_range: str
    section_references: list[str]
    section_calculations: list[dict[str, Any]]
    text: str
    js_hint: str


SECTION_COLS = {"B", "C", "G", "Y", "X"}
REFERENCE_COLS = {"T", "P", "Q", "R", "S"}
METHOD_KEYWORDS = (
    "fp innovations",
    "hamm",
    "en 1995",
    "usa clt",
    "vibration design",
)


def stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_cell_value(ws, row: int, col: int) -> Any:
    if row < 1 or col < 1:
        return None
    return ws.cell(row=row, column=col).value


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_plain_text(value: Any) -> bool:
    return isinstance(value, str) and value.strip() and not value.strip().startswith("=")


def row_values(ws, row: int) -> list[tuple[int, Any]]:
    values: list[tuple[int, Any]] = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v not in (None, ""):
            values.append((col, v))
    return values


def row_has_formula(ws, row: int) -> bool:
    return any(is_formula(v) for _, v in row_values(ws, row))


def find_nearby_label(ws_formula, row: int, col: int, max_left: int = 6, max_up: int = 4) -> str:
    """Find a human label around a cell.

    Engineering spreadsheets often put labels to the left or above a formula/value.
    This heuristic prioritizes left cells, then cells above.
    """
    candidates: list[str] = []

    for offset in range(1, max_left + 1):
        v = get_cell_value(ws_formula, row, col - offset)
        if is_plain_text(v):
            candidates.append(v.strip())
            break

    for offset in range(1, max_up + 1):
        v = get_cell_value(ws_formula, row - offset, col)
        if is_plain_text(v):
            candidates.append(v.strip())
            break

    return " | ".join(candidates)


def get_nearby_context(ws_formula, ws_values, row: int, col: int, radius: int = 1) -> dict[str, Any]:
    context: dict[str, Any] = {}
    for r in range(max(1, row - radius), min(ws_formula.max_row, row + radius) + 1):
        for c in range(max(1, col - radius), min(ws_formula.max_column, col + radius) + 1):
            addr = f"{get_column_letter(c)}{r}"
            formula_value = ws_formula.cell(r, c).value
            computed_value = ws_values.cell(r, c).value
            if formula_value is not None or computed_value is not None:
                context[addr] = {
                    "formula_or_text": formula_value,
                    "computed_value": computed_value,
                }
    return context


def should_index_cell(value: Any, formula: str, label: str) -> bool:
    if formula:
        return True
    if label and value not in (None, ""):
        return True
    if isinstance(value, str) and len(value.strip()) >= 3:
        return True
    return False


def looks_like_heading(ws_formula, row: int) -> bool:
    """Detect section headings in engineering spreadsheets.

    The Vibration sheet uses text rows such as:
    - Calculate the root mean square acceleration
    - Calculate Effective Width
    - Floating Floor or Screed Properties

    A heading is usually a plain-text cell in B/C/G/X/Y and the row itself does
    not contain formulas. Ordinary calculation rows are excluded because they
    normally have formulas in D/E/etc.
    """
    if row_has_formula(ws_formula, row):
        return False

    for col, value in row_values(ws_formula, row):
        col_letter = get_column_letter(col)
        text = stringify(value)
        if not text or len(text) < 3:
            continue
        if col_letter not in SECTION_COLS:
            continue
        # Ignore short variable/field labels that happen to be alone.
        if len(text) <= 6 and not any(k in text.lower() for k in METHOD_KEYWORDS):
            continue
        return True
    return False


def heading_text(ws_formula, row: int) -> str:
    texts: list[str] = []
    for col, value in row_values(ws_formula, row):
        col_letter = get_column_letter(col)
        text = stringify(value)
        if is_plain_text(value) and col_letter in SECTION_COLS and len(text) >= 3:
            texts.append(text)
    return " | ".join(texts)


def is_major_heading(text: str) -> bool:
    lower = text.lower()
    return any(k in lower for k in METHOD_KEYWORDS)


def collect_section_references(ws_formula, start_row: int, end_row: int) -> list[str]:
    refs: list[str] = []
    # References are often placed just above the section title in purple callout rows.
    search_start = max(1, start_row - 5)
    for r in range(search_start, end_row + 1):
        for c in range(1, ws_formula.max_column + 1):
            col_letter = get_column_letter(c)
            v = ws_formula.cell(r, c).value
            if is_plain_text(v) and (col_letter in REFERENCE_COLS or "equation" in v.lower() or "clause" in v.lower() or "section" in v.lower()):
                text = stringify(v)
                if text and text not in refs:
                    refs.append(text)
    return refs


def collect_section_calculations(
    ws_formula,
    ws_values,
    start_row: int,
    end_row: int,
    *,
    max_items: int = 40,
) -> list[dict[str, Any]]:
    """Collect calculation rows inside one visual/logical section.

    This is the key addition for RAG: formula cells are no longer isolated.
    A cell like D451 (arms) will carry the surrounding chain from M*, kres,
    damping factor, Fdyn, etc., so Gemini can explain the full section.
    """
    items: list[dict[str, Any]] = []
    for r in range(start_row, end_row + 1):
        row_label = ""
        # Prefer B/C labels, fallback to left-most plain text.
        for c in (2, 3, 6, 7):
            v = ws_formula.cell(r, c).value
            if is_plain_text(v):
                row_label = stringify(v)
                break
        if not row_label:
            for c, v in row_values(ws_formula, r):
                if is_plain_text(v):
                    row_label = stringify(v)
                    break

        formula_cells: list[dict[str, Any]] = []
        value_cells: list[dict[str, Any]] = []
        notes: list[str] = []
        for c, v in row_values(ws_formula, r):
            addr = f"{get_column_letter(c)}{r}"
            computed = ws_values.cell(r, c).value
            if is_formula(v):
                formula_cells.append({
                    "address": addr,
                    "formula": normalize_formula(v),
                    "value": computed,
                    "dependencies": extract_references(v),
                })
            elif is_plain_text(v):
                col_letter = get_column_letter(c)
                if col_letter not in {"B", "C"} and stringify(v) != row_label:
                    notes.append(stringify(v))
            elif v is not None and c <= 6:
                value_cells.append({"address": addr, "value": computed})

        if row_label or formula_cells:
            items.append({
                "row": r,
                "label": row_label,
                "formulas": formula_cells,
                "values": value_cells,
                "notes": notes[:4],
            })
        if len(items) >= max_items:
            break
    return items


def build_section_map(ws_formula, ws_values) -> dict[int, dict[str, Any]]:
    """Map every row to the nearest logical calculation section.

    The map is based on heading rows, not fixed row numbers, so it survives
    workbook changes better than a hard-coded list.
    """
    headings: list[tuple[int, str]] = []
    for row in range(1, ws_formula.max_row + 1):
        if looks_like_heading(ws_formula, row):
            title = heading_text(ws_formula, row)
            if title:
                headings.append((row, title))

    # Fallback: if no headings are found, treat the whole sheet as one section.
    if not headings:
        return {
            row: {
                "title": ws_formula.title,
                "path": [ws_formula.title],
                "start_row": 1,
                "end_row": ws_formula.max_row,
                "range": f"1:{ws_formula.max_row}",
                "references": [],
                "calculations": [],
            }
            for row in range(1, ws_formula.max_row + 1)
        }

    sections: list[dict[str, Any]] = []
    current_major = ""
    for i, (start, title) in enumerate(headings):
        end = headings[i + 1][0] - 1 if i + 1 < len(headings) else ws_formula.max_row
        if is_major_heading(title):
            current_major = title
        path = [p for p in [current_major, title] if p]
        # Avoid duplicate path when the heading itself is the major heading.
        if len(path) == 2 and path[0] == path[1]:
            path = [title]
        sections.append({
            "title": title,
            "path": path,
            "start_row": start,
            "end_row": end,
            "range": f"{start}:{end}",
            "references": collect_section_references(ws_formula, start, end),
            "calculations": collect_section_calculations(ws_formula, ws_values, start, end),
        })

    row_to_section: dict[int, dict[str, Any]] = {}
    for section in sections:
        for row in range(section["start_row"], section["end_row"] + 1):
            row_to_section[row] = section
    return row_to_section


def build_text(doc: dict[str, Any]) -> str:
    pieces = [
        f"Workbook: {doc['workbook']}",
        f"Sheet: {doc['sheet']}",
        f"Cell: {doc['address']}",
        f"Section path: {' > '.join(doc.get('section_path', []))}",
        f"Section title: {doc.get('section_title', '')}",
        f"Section rows: {doc.get('section_range', '')}",
        f"Section references: {' | '.join(doc.get('section_references', []))}",
        f"Label: {doc['label']}",
        f"Value: {doc['value']}",
        f"Formula: {doc['formula']}",
        f"Variable name: {doc.get('variable_name', '')}",
        f"Variable role: {doc.get('variable_role', '')}",
        f"Dependencies: {', '.join(doc['dependencies'])}",
        f"Named dependencies: {doc.get('named_dependencies', {})}",
        f"Named JavaScript formula: {doc.get('js_formula_named', '')}",
        f"Risk: {doc['risk_level']} {' '.join(doc['risk_reasons'])}",
    ]

    section_calculations = doc.get("section_calculations", [])
    if section_calculations:
        pieces.append("Section calculation chain:")
        for item in section_calculations:
            label = item.get("label", "")
            formulas = item.get("formulas", [])
            if formulas:
                for f in formulas:
                    pieces.append(
                        f"- Row {item.get('row')} {label}: {f.get('address')} = {f.get('formula')} -> {f.get('value')}"
                    )
            elif label:
                pieces.append(f"- Row {item.get('row')}: {label}")

    for addr, item in doc["nearby_context"].items():
        pieces.append(f"Nearby {addr}: {item}")
    return "\n".join(p for p in pieces if p and p != "None")


def extract_sheet_documents(
    workbook_path: str | Path,
    sheet_name: str = "Vibration",
    output_jsonl: str | Path | None = None,
    only_formulas: bool = False,
    include_section_context: bool = True,
) -> list[CellDocument]:
    workbook_path = Path(workbook_path)
    wb_formula = load_workbook(workbook_path, data_only=False, read_only=False)
    wb_values = load_workbook(workbook_path, data_only=True, read_only=False)

    if sheet_name not in wb_formula.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb_formula.sheetnames}")

    ws_formula = wb_formula[sheet_name]
    ws_values = wb_values[sheet_name]
    section_map = build_section_map(ws_formula, ws_values) if include_section_context else {}
    variable_registry = build_variable_registry(ws_formula, ws_values, section_map, find_nearby_label)

    docs: list[CellDocument] = []
    for row in range(1, ws_formula.max_row + 1):
        for col in range(1, ws_formula.max_column + 1):
            cell = ws_formula.cell(row, col)
            value_cell = ws_values.cell(row, col)
            raw = cell.value
            formula = raw if is_formula(raw) else ""
            value = value_cell.value
            label = find_nearby_label(ws_formula, row, col)

            if only_formulas and not formula:
                continue
            if not should_index_cell(value, formula, label):
                continue

            address = f"{get_column_letter(col)}{row}"
            dependencies = extract_references(formula)
            variable_info = variable_registry.get(address.upper())
            variable_name = variable_info.name if variable_info else ""
            variable_role = variable_info.role if variable_info else "unknown"
            dep_names = named_dependencies(dependencies, variable_registry)
            js_formula_named = formula_to_named_js(formula, dep_names)
            risk = detect_formula_risk(formula)
            nearby_context = get_nearby_context(ws_formula, ws_values, row, col, radius=1)
            section = section_map.get(row, {})

            temp = {
                "workbook": workbook_path.name,
                "sheet": sheet_name,
                "address": address,
                "label": label,
                "value": value,
                "formula": normalize_formula(formula),
                "dependencies": dependencies,
                "variable_name": variable_name,
                "variable_role": variable_role,
                "named_dependencies": dep_names,
                "js_formula_named": js_formula_named,
                "risk_level": risk.level,
                "risk_reasons": risk.reasons,
                "nearby_context": nearby_context,
                "section_title": section.get("title", ""),
                "section_path": section.get("path", []),
                "section_range": section.get("range", ""),
                "section_references": section.get("references", []),
                "section_calculations": section.get("calculations", []),
            }

            doc = CellDocument(
                id=f"{workbook_path.stem}::{sheet_name}!{address}",
                workbook=workbook_path.name,
                sheet=sheet_name,
                address=address,
                row=row,
                column=col,
                label=label,
                value=value,
                formula=normalize_formula(formula),
                number_format=cell.number_format,
                dependencies=dependencies,
                variable_name=variable_name,
                variable_role=variable_role,
                named_dependencies=dep_names,
                js_formula_named=js_formula_named,
                risk_level=risk.level,
                risk_reasons=risk.reasons,
                nearby_context=nearby_context,
                section_title=temp["section_title"],
                section_path=temp["section_path"],
                section_range=temp["section_range"],
                section_references=temp["section_references"],
                section_calculations=temp["section_calculations"],
                text=build_text(temp),
                js_hint=excel_formula_to_js_hint(formula),
            )
            docs.append(doc)

    if output_jsonl:
        output_path = Path(output_jsonl)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8") as f:
            for doc in docs:
                f.write(json.dumps(asdict(doc), ensure_ascii=False, default=str) + "\n")

    return docs
