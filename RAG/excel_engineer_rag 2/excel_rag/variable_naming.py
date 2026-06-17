from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl.utils import get_column_letter

from .formula_utils import excel_formula_to_js_hint

WORD_RE = re.compile(r"[A-Za-z]+\d*|\d+")
CELL_REF_RE = re.compile(
    r"(?:(?:'[^']+'|[A-Za-z0-9_ ]+)!)?\$?[A-Z]{1,3}\$?\d+", re.IGNORECASE
)

GREEK_MAP = {
    "ζ": "zeta",
    "ξ": "xi",
    "μ": "mu",
    "ρ": "rho",
    "π": "pi",
    "Σ": "sum",
    "∆": "delta",
    "Δ": "delta",
}

# Common compact engineering symbols used in vibration / timber design sheets.
SYMBOL_ALIASES = {
    "arms": "aRms",
    "a rms": "aRms",
    "a_rms": "aRms",
    "ar.m.s": "aRms",
    "vrms": "vRms",
    "v rms": "vRms",
    "v_rms": "vRms",
    "vmax": "vMax",
    "v max": "vMax",
    "f1": "f1",
    "fn": "naturalFrequency",
    "kres": "kRes",
    "k res": "kRes",
    "fdyn": "fDyn",
    "f dyn": "fDyn",
    "μres": "muRes",
    "mures": "muRes",
    "mu res": "muRes",
    "zeta damping factor": "dampingFactor",
    "zeta dumping factor": "dampingFactor",
    "damping factor": "dampingFactor",
    "dumping factor": "dampingFactor",
    "armslim": "aRmsLimit",
    "arms lim": "aRmsLimit",
    "mstar": "modalMass",
    "m*": "modalMass",
    "ei": "bendingStiffness",
    "e i": "bendingStiffness",
}

STOP_WORDS = {
    "the",
    "of",
    "for",
    "to",
    "in",
    "on",
    "and",
    "or",
    "with",
    "by",
    "from",
    "calculate",
    "calculated",
    "calculation",
    "check",
    "value",
    "input",
    "output",
    "mm",
    "m",
    "n",
    "kn",
    "kg",
    "pa",
    "mpa",
    "gpa",
    "hz",
    "sec",
    "s",
    "row",
    "cell",
    "formula",
    "result",
}

ABBREVIATION_WORDS = {
    "rms": "Rms",
    "clt": "Clt",
    "lvl": "Lvl",
    "glulam": "Glulam",
    "en": "En",
    "usa": "Usa",
    "uk": "Uk",
    "ei": "EI",
    "frl": "Frl",
}


@dataclass(frozen=True)
class VariableInfo:
    name: str
    role: str
    source_label: str
    source_text: str
    address: str


def stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(text: str) -> str:
    text = stringify(text)
    for old, new in GREEK_MAP.items():
        text = text.replace(old, f" {new} ")
    text = text.replace("²", "2").replace("³", "3").replace("⁴", "4")
    text = text.replace("/", " ").replace("\\", " ")
    text = re.sub(r"\([^)]*\)", " ", text)  # remove unit annotations in brackets
    text = re.sub(r"\[[^]]*\]", " ", text)
    text = re.sub(r"[^A-Za-z0-9_*]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def split_words(text: str) -> list[str]:
    normalized = normalize_text(text)
    return [
        w.lower() for w in WORD_RE.findall(normalized) if w.lower() not in STOP_WORDS
    ]


def word_to_pascal(word: str) -> str:
    lower = word.lower()
    if lower in ABBREVIATION_WORDS:
        return ABBREVIATION_WORDS[lower]
    if lower.isdigit():
        return lower
    return lower[:1].upper() + lower[1:]


def to_camel_case(text: str, *, fallback: str = "value") -> str:
    raw = normalize_text(text).lower()
    if raw in SYMBOL_ALIASES:
        return SYMBOL_ALIASES[raw]

    words = split_words(text)
    if not words:
        words = split_words(fallback)
    if not words:
        return "value"

    first = words[0]
    if first in SYMBOL_ALIASES:
        base = SYMBOL_ALIASES[first]
        rest = words[1:]
        return base + "".join(word_to_pascal(w) for w in rest)

    first_out = first.lower()
    # Keep common leading engineering symbols readable.
    if first_out == "rms":
        first_out = "rms"
    return first_out + "".join(word_to_pascal(w) for w in words[1:])


def to_pascal_case(text: str, *, fallback: str = "Value") -> str:
    camel = to_camel_case(text, fallback=fallback)
    return camel[:1].upper() + camel[1:] if camel else "Value"


def infer_variable_role(formula: str, value: Any, label: str) -> str:
    """Classify names for output conventions.

    - formula cells become calculated variables and use camelCase.
    - non-formula scalar inputs become constants and use PascalCase/Capitalized.
    """
    if formula:
        return "calculated"
    if value not in (None, ""):
        return "constant"
    if label:
        return "label"
    return "unknown"


def choose_source_text(label: str, section_title: str, address: str) -> str:
    label = stringify(label)
    section_title = stringify(section_title)

    # Very short symbolic labels are often clearer with section context.
    if label and len(normalize_text(label)) <= 4 and section_title:
        # Preserve well-known aliases such as arms/kres, otherwise enrich with section.
        if normalize_text(label).lower() in SYMBOL_ALIASES:
            return label
        return f"{label} {section_title}"
    if label:
        return label
    if section_title:
        return f"{section_title} {address}"
    return f"cell {address}"


def make_unique(name: str, used: dict[str, int]) -> str:
    if name not in used:
        used[name] = 1
        return name
    used[name] += 1
    return f"{name}{used[name]}"


def build_variable_registry(
    ws_formula, ws_values, section_map: dict[int, dict[str, Any]], find_label_fn
) -> dict[str, VariableInfo]:
    """Create a stable cell -> variable-name registry for the whole sheet.

    This runs before document extraction so dependency cells can be named even
    when --only-formulas is used. That means a formula doc can still say:
    D451 = (kRes * muRes * fDyn) / (...), instead of D451 = (D447*D450*D449)/(...).
    """
    used: dict[str, int] = {}
    registry: dict[str, VariableInfo] = {}

    for row in range(1, ws_formula.max_row + 1):
        section = section_map.get(row, {})
        section_title = stringify(section.get("title", ""))
        for col in range(1, ws_formula.max_column + 1):
            address = f"{get_column_letter(col)}{row}"
            raw = ws_formula.cell(row, col).value
            value = ws_values.cell(row, col).value
            formula = raw if isinstance(raw, str) and raw.startswith("=") else ""
            label = find_label_fn(ws_formula, row, col)
            plain_text = (
                raw
                if isinstance(raw, str)
                and raw.strip()
                and not raw.strip().startswith("=")
                else ""
            )

            # Only name meaningful cells: formulas, numeric constants, or labelled values.
            if not formula and value in (None, "") and not label:
                continue
            if plain_text and not label and value == raw:
                # Headings/notes do not need variable names.
                continue

            role = infer_variable_role(formula, value, label)
            source_text = choose_source_text(label, section_title, address)
            if role == "constant":
                candidate = to_pascal_case(source_text, fallback=address)
            else:
                candidate = to_camel_case(source_text, fallback=address)
            candidate = make_unique(candidate, used)
            registry[address.upper()] = VariableInfo(
                name=candidate,
                role=role,
                source_label=label,
                source_text=source_text,
                address=address,
            )
    return registry


def named_dependencies(
    dependencies: list[str], registry: dict[str, VariableInfo]
) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for dep in dependencies:
        key = dep.replace("$", "").upper()
        # Ranges and external refs are kept as-is for now.
        if ":" in key or "!" in key:
            out[dep] = {"name": dep, "role": "range_or_external", "label": ""}
            continue
        info = registry.get(key)
        if info:
            out[dep] = {
                "name": info.name,
                "role": info.role,
                "label": info.source_label,
            }
        else:
            out[dep] = {"name": dep.replace("$", ""), "role": "unknown", "label": ""}
    return out


def formula_to_named_js(formula: str, dep_names: dict[str, dict[str, str]]) -> str:
    if not formula:
        return ""
    js = excel_formula_to_js_hint(formula)

    # Replace longer refs first to avoid partial replacement.
    refs = sorted(dep_names.keys(), key=len, reverse=True)
    for ref in refs:
        if ":" in ref or "!" in ref:
            continue
        name = dep_names[ref]["name"]
        pattern = re.compile(
            rf"(?<![A-Za-z0-9_])\$?{re.escape(ref).replace('\\$', '\\$?')}(?![A-Za-z0-9_])",
            re.IGNORECASE,
        )
        js = pattern.sub(name, js)
    return js
